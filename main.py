# -*- coding: utf-8 -*-
from PIL import Image, ImageDraw, ImageFont
import textwrap
import openpyxl
import shutil
import os
from os import path
import traceback
import sys
import binascii
from msvcrt import getch

print('start')

# # Настройки
# Файл шрифта
fount_file = 'segoe_ui.ttf'

# Максимальная длина строки текста
max_width = 36

# Номер колонки с сылкой на картинку
img_col = 2

# Номер колонки с текстом
txt_col = 3


# def stop(symbol, message):
#     while True:
#         print(message)
#         if getch() == symbol:
#             break


# # загружаем файл excel
# file_xlsx = r"\\atlas\Общая\17.ИТ Дирекция\17.21 Ренкас Андрей\прайс_whatsapp3.xlsx"

if len(sys.argv) < 2:
    print("Не переданны параметры")
    exit(-1)

file_xlsx = sys.argv[1]

# move_to = r'\\atlas\Общая\17.ИТ Дирекция\17.21 Ренкас Андрей\img'
move_to = sys.argv[2]

# # Дебаг
# debag_value = False
# if len(sys.argv) == 3:
#     debag = sys.argv[3]
#     if debag == '-d':
#         debag_value = True


# print(file_xlsx)
# print(move_to)

# Парсим файл excel

try:
    wb_obj = openpyxl.load_workbook(file_xlsx)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
except Exception as e:
    print('Ошибка:\n', traceback.format_exc())
    # stop(b'\r', 'Press Enter to exit')

# Ошибка
error = False

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=img_col)
    try:
        # Определяем расширенее файла
        extension = cell_obj.value.split(".")
        extension = extension[-1]
        file = move_to + "\\" + 'temp_image.' + extension

        # if debag_value == True:
        # print('--------------------------------------')
        # print('Файл картинки: ' + cell_obj.value)
        if cell_obj.value == "Фото":
            continue
        shutil.copyfile(cell_obj.value, file)

        # with open(cell_obj.value, "rb") as file_buf:
        #     file_img = file_buf.read()
        # with open(file, "wb+") as file_temp:
        #     file_temp.write(file_img)

        # Вычисляем хеш картинки
        hash_value = str(binascii.crc32(bytes(cell_obj.value, 'utf-8')))

        # hash_value = hash(cell_obj.value)
        # hash_value = str(hash_value)
        # hash_value = hash_value.replace("-", "")

        # Картинка
        img = Image.open(file)
        # получаем ширину и высоту
        width, height = img.size
        # print(width, height)

        # Сохраняем картинку учитывая соотношение сторон
        # Ширина картинки
        fixed_width = 600

        # получаем процентное соотношение
        # старой и новой ширины
        width_percent = (fixed_width / float(img.size[0]))

        # на основе предыдущего значения
        # вычисляем новую высоту
        height_size = int((float(img.size[1]) * float(width_percent)))

        # меняем размер на полученные значения
        new_image = img.resize((fixed_width, height_size))
        new_image.save(file)

        # Преобразовываем текст
        # message = text_post
        cell_obj = sheet_obj.cell(row=i, column=txt_col)
        message = cell_obj.value
        message_arr = message.split('\n')
        # print(message_arr)

        message_txt = ''
        for message_item in message_arr:
            message_item_len = len(message_item)
            if message_item_len > max_width:
                message_item = textwrap.fill(message_item, width=max_width, initial_indent='', subsequent_indent='', expand_tabs=False, replace_whitespace=False, fix_sentence_endings=False, break_long_words=True, drop_whitespace=False, break_on_hyphens=False)
                message_txt = message_txt + message_item + '\n'
            else:
                message_txt = message_txt + message_item + '\n'

        # print(message_txt)

        old_img = Image.open(file)
        # создание нового изображения
        new_white = old_img.size[0] + fixed_width
        new_image = Image.new(old_img.mode, (round(new_white), old_img.size[1]), 'white')
        # вставляем старое изображение в новое
        new_image.paste(old_img, (0, 0))
        new_image.save(file)

        # Добавляем надпись
        img = Image.open(file)
        d1 = ImageDraw.Draw(img)
        path_ttf = os.path.abspath(__file__) + fount_file
        path_ttf = path_ttf.replace(r'main.py', r'')

        myFont = ImageFont.truetype(path_ttf, 30)
        d1.text((640, 40), message_txt, fill=(0, 0, 0), font=myFont)

        # Сохраняем и перемещаем файл
        file_name = message.split('\n')[1] + '.' + extension
        if path.exists(move_to):
            try:
                file_name = file_name.replace("/", " ")
                file_name = file_name.replace(r"\\", " ")
                file_name = hash_value + ' ' + file_name

                print('Сохраняем: ' + move_to + "\\" + file_name)
                img.save(move_to + "\\" + file_name)
                os.remove(file)
            except Exception as e:
                print('Ошибка:\n', traceback.format_exc())
                # stop(b'\r', 'Press Enter to exit')
                error = True
            finally:
                # print('Сохранен файл: ' + file_name)
                pass
        else:
            print('Не удалось подключиться к каталогу')
            # print(move_to + file_name)
            error = True
    # except FileNotFoundError:
    #     print('Не удалось скачать файл')
    except AttributeError:
        print('Нет файла картинки')
        # error = True
    except FileNotFoundError:
        print("Не удалось получить файл картинки " + message.split('\n')[2] + " " + message.split('\n')[1])
    except Exception as e:
        print('Ошибка:\n', traceback.format_exc())
        # stop(b'\r', 'Press Enter to exit')

if (error == True):
    # print(0)
    exit(-1)